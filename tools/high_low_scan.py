# -*- coding: utf-8 -*-
"""
미국·홍콩·중국A 시총 $10B+ 종목의 60일/52주 신고가·신저가 스캔 → xlsx + HTML 대시보드

사용법:
  python high_low_scan.py                          # 스캔 실행 → xlsx + scan csv + top_movers json
  python high_low_scan.py --merge-reasons R.json   # 당일 scan csv에 이유(json: {티커: 이유})를 합쳐 xlsx 재생성
  python high_low_scan.py --out-dir <폴더>          # 출력 폴더 지정 (기본: ~/Desktop/신고신저가)
  python high_low_scan.py --source legacy          # TradingView 대신 eastmoney/yfinance 경로 강제

데이터 소스 (무료):
  - 기본: TradingView scanner — 시장당 요청 1번에 유니버스·섹터·시세·3M/52주 고저·PER(12MF)·PBR까지.
  - 폴백(legacy): 유니버스·섹터·PER/PBR = NASDAQ screener(미국) + eastmoney clist(홍콩·중국A),
    일봉 = 전시장 yfinance. (eastmoney kline은 동시/연속 요청 IP차단이 심해 쓰지 않음.
    clist도 push2 본 호스트 차단이 잦아 push2delay 사용 + 직전 캐시 폴백.)
    미국 PER/PBR은 legacy에선 히트 종목만 야후 info로 사후 수집.
티커는 블룸버그 형식 + Equity 접미사 (AAPL US Equity / 700 HK Equity / 600519 CH Equity).
신고·신저 판정: 당일 고가 >= 직전 60일(TV는 3M≈63거래일)/52주 고가 최대치, 저가는 반대.
NEW/OLD: 직전 scan_*.csv 대비 신규 진입이면 NEW, 연속 등재면 OLD(전일 이유 자동 승계).
섹터동향 시트: 시장×섹터 신고-신저 순강도로 강세/약세 판정 (유니버스 종목수 병기).
"""
import argparse
import concurrent.futures as cf
import datetime as dt
import json
import os
import re
import sys
import time
from pathlib import Path

import numpy as np
import pandas as pd
import requests

sys.stdout.reconfigure(encoding="utf-8")

MCAP_USD = 10e9
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)", "Accept": "application/json"}
EM_HEADERS = {"User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"}
FX_FALLBACK = {"HKD": 7.8, "CNY": 7.2, "JPY": 145.0}
TOP_N = 5  # 시장·방향별 이유 붙일 신규(NEW) 상위 종목 수
OLD_EVENT_CHG = 7.0  # OLD여도 당일 |등락률|이 이 값 이상이면 새 이벤트로 보고 이유 재검색
OLD_TOP_N = 3        # 시장·방향별 OLD 재검색 상한(급변은 드무니 소수만)

COLS = ["구분", "NEW/OLD", "티커", "종목명", "섹터", "종가", "등락률(%)", "시총($B)",
        "PER(12MF)", "PER(TTM)", "PBR", "60일", "52주", "이유"]
MARKET_ORDER = ["미국", "일본", "홍콩", "중국A"]


def log(msg):
    print(f"[{dt.datetime.now():%H:%M:%S}] {msg}", flush=True)


def fmt_val(x):
    """PER/PBR 표시용 — 양수만, 소수 1자리. 적자/결측은 None(빈칸)."""
    try:
        x = float(x)
    except (TypeError, ValueError):
        return None
    return round(x, 1) if x > 0 else None


# ---------------------------------------------------------------- FX
def get_fx():
    fx = dict(FX_FALLBACK)
    try:
        import yfinance as yf
        px = yf.download(["HKD=X", "CNY=X", "JPY=X"], period="5d",
                         progress=False, auto_adjust=True)["Close"]
        for cur, tk in (("HKD", "HKD=X"), ("CNY", "CNY=X"), ("JPY", "JPY=X")):
            v = px[tk].dropna()
            if len(v):
                fx[cur] = float(v.iloc[-1])
    except Exception as e:
        log(f"FX 조회 실패, 기본값 사용: {e}")
    log(f"환율 USDHKD={fx['HKD']:.3f} USDCNY={fx['CNY']:.3f} USDJPY={fx['JPY']:.1f}")
    return fx


# ---------------------------------------------------------------- 유니버스 (legacy)
def clean_us_name(name):
    return re.sub(
        r"\s+(Common Stock|Common Shares|Ordinary Shares|American Depositary Shares.*|ADS.*|"
        r"Class [A-C]( Common Stock| Ordinary Shares)?|\(.*\))\s*$", "", name).strip().rstrip(",")


def us_universe():
    r = requests.get("https://api.nasdaq.com/api/screener/stocks",
                     params={"tableonly": "true", "limit": "25", "offset": "0", "download": "true"},
                     headers=UA, timeout=60)
    df = pd.DataFrame(r.json()["data"]["rows"])
    df["mcap"] = pd.to_numeric(df["marketCap"], errors="coerce")
    df = df[df["mcap"] >= MCAP_USD]
    df = df[~df["symbol"].str.contains(r"[\^/ ]")]
    df = df[~df["name"].str.contains(r"ETF|Preferred|Warrant|Unit(?:s)?\b", case=False, regex=True)]
    out = pd.DataFrame({
        "market": "미국",
        "yahoo": df["symbol"].str.replace(".", "-", regex=False),
        "bb": df["symbol"].str.replace(".", "/", regex=False) + " US Equity",
        "name": df["name"].map(clean_us_name),
        "sector": df["sector"].replace("", "-").fillna("-"),
        "mcap_usd": df["mcap"],
        "pe_ttm": np.nan,  # 미국 legacy는 히트 종목만 야후에서 사후 수집
        "pbr": np.nan,
    })
    return out.reset_index(drop=True)


def em_list(fs, mcap_floor_local):
    """eastmoney 종목 리스트 — 시총 내림차순 페이지를 임계값 아래로 내려갈 때까지 수집."""
    rows, pn = [], 1
    # push2 본 호스트는 연속요청 시 IP 차단이 걸림 → 지연시세 호스트(push2delay)를 기본으로
    hosts = ["push2delay.eastmoney.com", "push2.eastmoney.com"]
    backoffs = [15, 30, 60, 120, 240]
    while True:
        for attempt in range(len(backoffs) + 1):
            try:
                r = requests.get(f"https://{hosts[attempt % len(hosts)]}/api/qt/clist/get",
                                 params={"pn": pn, "pz": 500, "po": 1, "np": 1, "fltt": 2,
                                         "invt": 2, "fid": "f20", "fs": fs,
                                         "fields": "f12,f13,f14,f20,f100,f115,f23"},
                                 headers=EM_HEADERS, timeout=30)
                break
            except requests.RequestException:
                if attempt == len(backoffs):
                    raise
                log(f"clist 실패, {backoffs[attempt]}초 대기 후 재시도…")
                time.sleep(backoffs[attempt])
        diff = (r.json().get("data") or {}).get("diff") or []
        if not diff:
            break
        for d in diff:
            if isinstance(d.get("f20"), (int, float)):
                rows.append(d)
        last = diff[-1].get("f20")
        if not isinstance(last, (int, float)) or last < mcap_floor_local:
            break
        pn += 1
        time.sleep(1.5)  # 페이지 간 간격 — 차단 예방
    df = pd.DataFrame(rows)
    df = df[pd.to_numeric(df["f20"], errors="coerce") >= mcap_floor_local]
    for c in ("f115", "f23"):
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def hk_universe(fx):
    df = em_list("m:128+t:3,m:128+t:4,m:128+t:1,m:128+t:2", MCAP_USD * fx["HKD"])
    return pd.DataFrame({
        "market": "홍콩",
        "yahoo": df["f12"].map(lambda c: f"{int(c):04d}.HK"),
        "bb": df["f12"].map(lambda c: str(int(c)) + " HK Equity"),
        "name": df["f14"],
        "sector": df["f100"].replace("", "-").fillna("-"),
        "mcap_usd": df["f20"] / fx["HKD"],
        "pe_ttm": df["f115"],
        "pbr": df["f23"],
    }).reset_index(drop=True)


def cn_yahoo(code, mkt):
    if int(mkt) == 1:
        return code + ".SS"
    if code[:2] in ("43", "83", "87", "92"):  # 북교소
        return code + ".BJ"
    return code + ".SZ"


def cn_universe(fx):
    df = em_list("m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23,m:0+t:81+s:2048", MCAP_USD * fx["CNY"])
    return pd.DataFrame({
        "market": "중국A",
        "yahoo": [cn_yahoo(c, m) for c, m in zip(df["f12"], df["f13"])],
        "bb": df["f12"] + " CH Equity",
        "name": df["f14"],
        "sector": df["f100"].replace("", "-").fillna("-"),
        "mcap_usd": df["f20"] / fx["CNY"],
        "pe_ttm": df["f115"],
        "pbr": df["f23"],
    }).reset_index(drop=True)


def get_universes(fx, cache_dir):
    """유니버스 3종 수집. 실패한 시장은 직전 성공 캐시로 폴백."""
    cache_dir.mkdir(parents=True, exist_ok=True)
    dfs = {}
    for name, fn in (("us", us_universe), ("hk", lambda: hk_universe(fx)),
                     ("cn", lambda: cn_universe(fx))):
        p = cache_dir / f"universe_{name}.csv"
        try:
            df = fn()
            df.to_csv(p, index=False, encoding="utf-8-sig")
        except Exception as e:
            if not p.exists():
                raise
            df = pd.read_csv(p, encoding="utf-8-sig", dtype=str)
            for c in ("mcap_usd", "pe_ttm", "pbr"):
                if c in df.columns:
                    df[c] = pd.to_numeric(df[c], errors="coerce")
            log(f"{name} 유니버스 수집 실패 → 캐시 사용: {type(e).__name__}")
        dfs[name] = df
    return dfs["us"], dfs["hk"], dfs["cn"]


# ---------------------------------------------------------------- 일봉 (legacy, 전시장 yfinance)
def fetch_history(symbols, label):
    import yfinance as yf
    out = {}
    CHUNK = 150
    for i in range(0, len(symbols), CHUNK):
        chunk = symbols[i:i + CHUNK]
        for attempt in range(3):
            try:
                data = yf.download(chunk, period="1y", interval="1d", auto_adjust=True,
                                   group_by="ticker", threads=True, progress=False)
                break
            except Exception as e:
                log(f"{label} yfinance chunk {i} 재시도 {attempt + 1}: {e}")
                time.sleep(5 * (attempt + 1))
        else:
            continue
        for sym in chunk:
            try:
                df = data[sym][["High", "Low", "Close"]].dropna(how="all")
            except KeyError:
                continue
            if len(df) >= 41:
                out[sym] = pd.DataFrame({"date": df.index.strftime("%Y-%m-%d"),
                                         "high": df["High"].values,
                                         "low": df["Low"].values,
                                         "close": df["Close"].values})
        log(f"{label} 일봉 {min(i + CHUNK, len(symbols))}/{len(symbols)}")
    return out


def fill_us_valuation(d):
    """미국 히트 종목만 야후 info에서 PER/PBR 수집 (legacy 경로용 — TV 경로는 스캐너가 줌)."""
    import yfinance as yf

    def one(bb):
        sym = bb.replace(" US Equity", "").replace("/", "-")
        try:
            info = yf.Ticker(sym).info
            pe = info.get("forwardPE") or info.get("trailingPE")
            return bb, fmt_val(pe), fmt_val(info.get("priceToBook"))
        except Exception:
            return bb, None, None

    with cf.ThreadPoolExecutor(8) as ex:
        res = {bb: (pe, pb) for bb, pe, pb in ex.map(one, list(d["티커"]))}
    d["PER(12MF)"] = [res[t][0] for t in d["티커"]]
    d["PBR"] = [res[t][1] for t in d["티커"]]


# ---------------------------------------------------------------- TradingView (기본 경로)
# (TV시장, 통화, 거래소) — is_primary 필터는 A+H 동시상장의 H주를 제외해버려서 쓰지 않고,
# 거래소+통화로 위안화 카운터(HK)·B주(중국)를 걸러낸다. ADR·듀얼클래스는 포함.
TV_MARKETS = {"미국": ("america", "USD", ["NASDAQ", "NYSE", "AMEX"]),
              "일본": ("japan", "JPY", ["TSE"]),
              "홍콩": ("hongkong", "HKD", ["HKEX"]),
              "중국A": ("china", "CNY", ["SSE", "SZSE"])}
TV_COLS = ["name", "description", "sector", "market_cap_basic", "close", "change",
           "high", "low", "price_52_week_high", "price_52_week_low", "High.3M", "Low.3M",
           "time", "currency", "price_earnings_fwd", "price_earnings_ttm", "price_book_fq"]
TV_NOTE = "60일 컬럼은 TradingView 3개월(약 63거래일) 고저가 기준. 신고/신저 판정: 당일 고가(저가)가 해당 기간 최고(최저)가 도달."
LEGACY_NOTE = ("60일 컬럼은 정확한 60거래일 고저가 기준 (eastmoney/yfinance 폴백 경로). "
               "주의: 폴백 경로의 PER은 12MF가 아니라 TTM 기준(소스가 선행 컨센서스를 주지 않음).")


def tv_market_scan(market, floor_local, exchanges, cur):
    body = {"filter": [
                {"left": "market_cap_basic", "operation": "egreater", "right": floor_local},
                {"left": "exchange", "operation": "in_range", "right": exchanges},
                {"left": "currency", "operation": "in_range", "right": [cur]},
                {"left": "typespecs", "operation": "has_none_of", "right": ["preferred"]}],
            "options": {"lang": "en"}, "markets": [market],
            "symbols": {"query": {"types": ["stock"]}, "tickers": []},
            "columns": TV_COLS,
            "sort": {"sortBy": "market_cap_basic", "sortOrder": "desc"},
            "range": [0, 6000]}
    r = requests.post(f"https://scanner.tradingview.com/{market}/scan",
                      json=body, headers=UA, timeout=60)
    r.raise_for_status()
    return [dict(zip(TV_COLS, row["d"])) for row in r.json().get("data", [])]


def scan_tradingview(fx, out_dir):
    per_market, data_dates, uni_sect = {}, {}, {}
    for m, (mk, cur, exchanges) in TV_MARKETS.items():
        rate = 1.0 if cur == "USD" else fx[cur]
        raw = tv_market_scan(mk, MCAP_USD * rate, exchanges, cur)
        if mk == "japan":
            # 일본은 legacy 유니버스 소스(나스닥/eastmoney 상당)가 없어 TV 성공분을 캐시 → 폴백용
            cache_dir = out_dir / "_cache"
            cache_dir.mkdir(parents=True, exist_ok=True)
            pd.DataFrame({
                "market": m,
                "yahoo": [str(d["name"]) + ".T" for d in raw],
                "bb": [str(d["name"]) + " JP Equity" for d in raw],
                "name": [d.get("description") or str(d["name"]) for d in raw],
                "sector": [d.get("sector") or "-" for d in raw],
                "mcap_usd": [d["market_cap_basic"] / rate for d in raw],
                "pe_fwd": [d.get("price_earnings_fwd") for d in raw],
                "pbr": [d.get("price_book_fq") for d in raw],
            }).to_csv(cache_dir / "universe_jp.csv", index=False, encoding="utf-8-sig")
        uni_sect[m] = {}
        rows, last_ts = [], 0
        for d in raw:
            s = d.get("sector") or "-"
            st = uni_sect[m].setdefault(s, {"n": 0, "pe": [], "pe_ttm": [], "pb": []})
            st["n"] += 1
            for k, f in (("pe", "price_earnings_fwd"), ("pe_ttm", "price_earnings_ttm"),
                         ("pb", "price_book_fq")):
                v = d.get(f)
                if isinstance(v, (int, float)) and v > 0:
                    st[k].append(v)
            hi, lo, h3, l3 = d.get("high"), d.get("low"), d.get("High.3M"), d.get("Low.3M")
            if None in (hi, lo, h3, l3):
                continue
            if d.get("time"):
                last_ts = max(last_ts, int(d["time"]))
            nh3, nl3 = hi >= h3, lo <= l3
            if not (nh3 or nl3):
                continue
            h52, l52 = d.get("price_52_week_high"), d.get("price_52_week_low")
            direction = "신고가" if nh3 else "신저가"
            sym = str(d["name"])
            bb = (sym.replace(".", "/") + " US" if mk == "america"
                  else str(int(sym)) + " HK" if mk == "hongkong"
                  else sym + " JP" if mk == "japan" else sym + " CH") + " Equity"
            chg = d.get("change")
            rows.append({"구분": direction, "NEW/OLD": "NEW", "티커": bb,
                         "종목명": d.get("description") or sym, "섹터": s,
                         "종가": round(float(d["close"]), 2) if d.get("close") is not None else np.nan,
                         "등락률(%)": round(float(chg), 2) if chg is not None else np.nan,
                         "시총($B)": round(d["market_cap_basic"] / rate / 1e9, 1),
                         "PER(12MF)": fmt_val(d.get("price_earnings_fwd")),
                         "PER(TTM)": fmt_val(d.get("price_earnings_ttm")),
                         "PBR": fmt_val(d.get("price_book_fq")),
                         "60일": "O",
                         "52주": "O" if ((h52 is not None and hi >= h52) or
                                         (l52 is not None and lo <= l52)) else "",
                         "이유": ""})
        per_market[m] = pd.DataFrame(rows, columns=COLS)
        data_dates[m] = (dt.datetime.fromtimestamp(last_ts, dt.timezone.utc).strftime("%Y-%m-%d")
                         if last_ts else dt.date.today().isoformat())
        log(f"TV {m}: 유니버스 {len(raw)}종목")
    return per_market, data_dates, uni_sect


# ---------------------------------------------------------------- 판정 (legacy)
def judge(df, market_last_date):
    """df: date/high/low/close 오름차순. 반환 None 또는 dict."""
    df = df.dropna(subset=["high", "low", "close"])
    if len(df) < 41 or df["date"].iloc[-1] != market_last_date:
        return None  # 데이터 부족 or 당일 미거래(정지)
    today, prior = df.iloc[-1], df.iloc[:-1]
    chg = (today["close"] / prior["close"].iloc[-1] - 1) * 100 if prior["close"].iloc[-1] else np.nan
    nh60 = today["high"] >= prior["high"].tail(60).max()
    nl60 = today["low"] <= prior["low"].tail(60).min()
    full52 = len(prior) >= 200  # 상장 1년 미만이면 52주 판정 생략
    nh52 = full52 and today["high"] >= prior["high"].tail(252).max()
    nl52 = full52 and today["low"] <= prior["low"].tail(252).min()
    if not (nh60 or nl60):
        return None
    return {"종가": round(float(today["close"]), 2), "등락률(%)": round(float(chg), 2),
            "nh60": bool(nh60), "nl60": bool(nl60), "nh52": bool(nh52), "nl52": bool(nl52)}


def build_rows(univ, hists):
    if not hists:
        return pd.DataFrame(columns=COLS), None
    last_date = max(df["date"].iloc[-1] for df in hists.values())
    rows = []
    for _, u in univ.iterrows():
        df = hists.get(u["yahoo"])
        if df is None:
            continue
        j = judge(df, last_date)
        if j is None:
            continue
        direction = "신고가" if j["nh60"] else "신저가"
        rows.append({"구분": direction, "NEW/OLD": "NEW",
                     "티커": u["bb"], "종목명": u["name"], "섹터": u["sector"],
                     "종가": j["종가"], "등락률(%)": j["등락률(%)"],
                     "시총($B)": round(u["mcap_usd"] / 1e9, 1),
                     "PER(12MF)": fmt_val(u.get("pe_fwd", u.get("pe_ttm"))),
                     "PER(TTM)": fmt_val(u.get("pe_ttm")),
                     "PBR": fmt_val(u.get("pbr")),
                     "60일": "O" if (j["nh60"] or j["nl60"]) else "",
                     "52주": "O" if (j["nh52"] or j["nl52"]) else "", "이유": ""})
    out = pd.DataFrame(rows, columns=COLS)
    if len(out):
        out["_s"] = np.where(out["구분"] == "신고가", -out["등락률(%)"], out["등락률(%)"])
        out = out.sort_values(["구분", "_s"], ascending=[False, True]).drop(columns="_s")
    return out.reset_index(drop=True), last_date


def load_prev_scan(out_dir, tag):
    """직전 스캔 csv 로드 — NEW/OLD 판정 + OLD 종목 이유 승계용."""
    files = sorted(p for p in out_dir.glob("scan_*.csv")
                   if p.stem.split("_")[-1].isdigit() and p.stem.split("_")[-1] < tag)
    if not files:
        return None
    return pd.read_csv(files[-1], encoding="utf-8-sig")


def norm_tk(t):
    """티커 정규화 — Equity 접미사 유무를 무시하고 비교 (구버전 csv 호환)."""
    return str(t).removesuffix(" Equity").strip()


def apply_new_old(per_market, prev):
    """전일 리스트와 비교해 NEW/OLD 태깅(OLD는 전일 이유 승계) 후
    신고가→신저가, NEW 먼저 순으로 재정렬."""
    if prev is not None:
        keys = set(zip(prev["시장"], prev["티커"].map(norm_tk), prev["구분"]))
        prev_reasons = {(r["시장"], norm_tk(r["티커"]), r["구분"]): r["이유"].strip()
                        for _, r in prev.iterrows()
                        if isinstance(r["이유"], str) and r["이유"].strip()}
    for m, d in per_market.items():
        if not len(d):
            continue
        if prev is None:
            d["NEW/OLD"] = "NEW"  # 첫 실행: 비교 기준 없음 → 전부 NEW
        else:
            tags, reasons = [], []
            for t, g in zip(d["티커"], d["구분"]):
                k = (m, norm_tk(t), g)
                if k in keys:
                    tags.append("OLD")
                    pr = prev_reasons.get(k, "")
                    reasons.append("(전일) " + pr.removeprefix("(전일) ") if pr else "")
                else:
                    tags.append("NEW")
                    reasons.append("")
            d["NEW/OLD"] = tags
            d["이유"] = reasons
        d["_s"] = np.where(d["구분"] == "신고가", -d["등락률(%)"], d["등락률(%)"])
        per_market[m] = (d.sort_values(["구분", "NEW/OLD", "_s"],
                                       ascending=[False, True, True])
                         .drop(columns="_s")[COLS].reset_index(drop=True))


# ---------------------------------------------------------------- 섹터 ETF 상대강도
SECTOR_ETFS = {
    # 순서 = 사용자 지정 고정순(엑셀 반출 시 이 순서 유지, 2026-07-21 확정)
    "미국": [("XLK", "테크"), ("XLV", "헬스케어"), ("XLF", "금융"), ("XLY", "경기소비재"),
            ("XLC", "커뮤니케이션"), ("XLI", "산업재"), ("XLP", "필수소비재"), ("XLE", "에너지"),
            ("XLU", "유틸리티"), ("XLB", "소재"), ("XLRE", "리츠")],
    "일본": [("1617.T", "식품"), ("1618.T", "에너지자원"), ("1619.T", "건설·자재"),
            ("1620.T", "소재·화학"), ("1621.T", "의약품"), ("1622.T", "자동차·운송기기"),
            ("1623.T", "철강·비철"), ("1624.T", "기계"), ("1625.T", "전기·정밀"),
            ("1626.T", "정보통신·서비스"), ("1627.T", "전력·가스"), ("1628.T", "운수·물류"),
            ("1629.T", "상사·도매"), ("1630.T", "소매"), ("1631.T", "은행"),
            ("1632.T", "금융(은행외)"), ("1633.T", "부동산")],
    "홍콩": [("2800.HK", "항셍지수"), ("2828.HK", "H주(중국기업)"), ("3033.HK", "항셍테크")],
    "중국A": [("512480.SS", "반도체"), ("515000.SS", "테크(科技)"), ("515050.SS", "5G통신"),
             ("515030.SS", "신에너지차"), ("512660.SS", "방산(军工)"), ("515790.SS", "태양광"),
             ("512690.SS", "백주(주류)"), ("512800.SS", "은행"), ("512400.SS", "유색금속"),
             ("512200.SS", "부동산"), ("512880.SS", "증권"), ("512010.SS", "의약"),
             ("159928.SZ", "소비(消费)")],
}
ETF_NOTE = ("미국=SPDR 섹터 ETF, 일본=TOPIX-17 섹터 ETF(1617~1633), 중국A=주요 섹터 ETF(상해·심천)"
            " · 홍콩은 섹터 ETF가 얇아 시장·테크 대표만"
            " · 수익률은 배당 포함 총수익 기준(미국=야후 수정주가, 중·홍=eastmoney 전복권)"
            " — 블룸버그 가격수익률과 고배당 섹터에서 연 1~2%p 차이 가능 · 정렬=시장·섹터 고정순(대시보드는 1D 등락순으로 표시)")


def etf_bb(sym):
    if sym.endswith(".HK"):
        return f"{int(sym[:-3])} HK"
    if sym.endswith((".SS", ".SZ")):
        return sym[:-3] + " CH"
    if sym.endswith(".T"):
        return sym[:-2] + " JP"
    return sym + " US"


def em_close_series(secid):
    """eastmoney 일봉 종가 시계열 (fqt=1 전복권) — A주 ETF 份额折算을 야후가 못 잡는 문제 회피."""
    backoffs = []  # TEMP(2026-07-24): eastmoney 차단 상태 — 빠른 실패로 전체 스캔 지연 회피. 복구 시 [3, 8, 20, 60]
    for attempt in range(len(backoffs) + 1):
        try:
            r = requests.get("https://push2his.eastmoney.com/api/qt/stock/kline/get",
                             params={"secid": secid, "klt": 101, "fqt": 1, "lmt": 1750,
                                     "end": "20500101", "fields1": "f1,f2",
                                     "fields2": "f51,f53"}, headers=EM_HEADERS, timeout=20)
            kl = (r.json().get("data") or {}).get("klines") or []
            if not kl:
                return pd.Series(dtype=float, index=pd.DatetimeIndex([]))
            rec = [x.split(",") for x in kl]
            return pd.Series([float(b) for _, b in rec],
                             index=pd.to_datetime([a for a, _ in rec])).dropna()
        except requests.RequestException:
            if attempt == len(backoffs):
                raise
            time.sleep(backoffs[attempt])


TX_URL = "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get"


def tx_sym(secid):
    """eastmoney secid → 텐센트 심볼 (1.=상해, 0.=심천, 116./100./124.=홍콩)."""
    mkt, _, code = secid.partition(".")
    return {"1": "sh", "0": "sz", "116": "hk", "100": "hk", "124": "hk"}.get(mkt, "") + code \
        if mkt in ("1", "0", "116", "100", "124") else None


def _tx_fetch(sym, scale, count, fq):
    """텐센트 kline 한 번 호출 → 종가 시계열. 레코드는 [날짜,시가,종가,고가,저가,거래량]."""
    r = requests.get(TX_URL, params={"param": f"{sym},{scale},,,{count},{fq}"},
                     headers={"User-Agent": "Mozilla/5.0"}, timeout=25)
    r.raise_for_status()
    node = (r.json().get("data") or {}).get(sym) or {}
    for k in (f"{fq}{scale}", scale):  # A주는 qfqday/qfqweek, 홍콩·지수는 복권이 없어 day
        rec = node.get(k)
        if isinstance(rec, list) and rec:
            return pd.Series([float(x[2]) for x in rec],
                             index=pd.to_datetime([x[0] for x in rec])).dropna()
    return pd.Series(dtype=float, index=pd.DatetimeIndex([]))


def tx_close_series(secid):
    """텐센트 일봉 종가 — eastmoney 차단 시 폴백(2026-07-31 추가).
    A주는 일봉 복권이 640행(2023-12~)까지만 와서, 그 이전 구간은 주봉 복권으로 이어붙인다
    (연도별 YTD는 연말 종가만 있으면 되므로 주봉 해상도로 충분)."""
    sym = tx_sym(secid)
    if not sym:
        return pd.Series(dtype=float, index=pd.DatetimeIndex([]))
    fq = "qfq" if sym[:2] in ("sh", "sz") else ""
    c = _tx_fetch(sym, "day", 2000, fq)
    if len(c) and c.index[0] > pd.Timestamp("2020-02-01"):
        w = _tx_fetch(sym, "week", 800, fq)
        if len(w):
            c = pd.concat([w[w.index < c.index[0]], c]).sort_index()
            c = c[~c.index.duplicated(keep="last")]
    return c


def cn_close_series(secid):
    """중국A·홍콩 일봉 종가 — eastmoney 우선, 실패하면 텐센트로 폴백.
    eastmoney는 2026-07-24부터 간헐 차단(성공률이 날마다 0~100%로 튐)이라 폴백이 필수."""
    try:
        c = em_close_series(secid)
        if len(c):
            return c
    except requests.RequestException:
        pass
    return tx_close_series(secid)


def etf_secid(sym):
    if sym.endswith(".HK"):
        return f"116.{int(sym[:-3]):05d}"  # eastmoney HK 코드는 5자리
    if sym.endswith(".SS"):
        return "1." + sym[:-3]
    if sym.endswith(".SZ"):
        return "0." + sym[:-3]
    return None  # 미국은 야후


def build_etf_table():
    """섹터별 대표 ETF의 단기(1D/1주/1M/3M) + 연도별 YTD 수익률·순위 테이블.
    미국·일본=야후(신뢰 OK), 홍콩·중국A=eastmoney(야후는 A주 ETF 折算 왜곡)."""
    import yfinance as yf
    yahoo_syms = [s for lst in SECTOR_ETFS.values() for s, _ in lst if etf_secid(s) is None]
    px = yf.download(yahoo_syms, start="2019-12-01", interval="1d", auto_adjust=True,
                     group_by="ticker", threads=True, progress=False)
    years = list(range(dt.date.today().year, 2019, -1))
    cols = ["시장", "티커", "섹터", "1D", "1주", "1M", "3M"]
    for y in years:
        cols += [f"YTD{y % 100}", f"순위{y % 100}"]
    recs = []
    for m, lst in SECTOR_ETFS.items():
        for sym, name in lst:
            sid = etf_secid(sym)
            try:
                if sid is None:
                    c = px[sym]["Close"].dropna()
                else:
                    c = cn_close_series(sid)
                    time.sleep(0.8)  # 순차 간격 — eastmoney 차단 예방
                c = c[c.index >= "2019-12-01"]
            except Exception as e:
                log(f"ETF {sym} 시세 실패(빈칸): {type(e).__name__}")
                c = pd.Series(dtype=float, index=pd.DatetimeIndex([]))
            r = {"시장": m, "티커": etf_bb(sym), "섹터": name}
            ret = lambda n: round(float(c.iloc[-1] / c.iloc[-1 - n] - 1) * 100, 1) if len(c) > n else None
            r["1D"], r["1주"], r["1M"], r["3M"] = ret(1), ret(5), ret(21), ret(63)
            for y in years:
                base = c[c.index.year == y - 1]  # 전년 말 종가 대비
                endp = c[c.index.year == y]
                r[f"YTD{y % 100}"] = (round(float(endp.iloc[-1] / base.iloc[-1] - 1) * 100, 1)
                                      if len(base) and len(endp) else None)
            recs.append(r)
    df = pd.DataFrame(recs)
    for y in years:
        df[f"순위{y % 100}"] = (df.groupby("시장")[f"YTD{y % 100}"]
                              .rank(ascending=False, method="min").astype("Int64"))
    df = df[cols]
    # 엑셀 반출용 고정 순서: 시장 순(MARKET_ORDER) + 시장 내 SECTOR_ETFS 정의순 유지.
    # (대시보드 표시는 renderEtf JS가 등락순으로 재정렬하므로 여기선 고정순만 보장)
    df["_m"] = pd.Categorical(df["시장"], categories=MARKET_ORDER, ordered=True)
    return (df.sort_values("_m", kind="stable")
            .drop(columns="_m").reset_index(drop=True))


# ---------------------------------------------------------------- 주요 지수
# (국가, 지수명, 소스, 코드) — 미·한·일=yfinance(신선), 중·홍=eastmoney(야후는 아시아지수 정체)
INDEX_DEFS = [
    ("미국", "S&P500", "y", "^GSPC"), ("미국", "나스닥종합", "y", "^IXIC"),
    ("한국", "KOSPI", "y", "^KS11"),
    ("일본", "닛케이225", "y", "^N225"),
    ("중국", "상해종합", "e", "1.000001"), ("중국", "CSI300", "e", "1.000300"),
    ("홍콩", "항셍", "e", "100.HSI"), ("홍콩", "항셍테크", "e", "124.HSTECH"),
]
INDEX_ORDER = ["미국", "한국", "일본", "중국", "홍콩"]
INDEX_COLS = ["국가", "지수", "지수레벨", "1D", "1주", "1M", "3M", "YTD", "52주위치%", "기준일"]


def build_index_table():
    """주요 지수 스냅샷 — 레벨·1D·1주·1M·3M·YTD·52주 레인지 내 위치·데이터 기준일.
    미·한·일=yfinance, 중·홍=eastmoney(아시아 지수는 야후가 며칠씩 정체돼 부정확)."""
    import yfinance as yf
    ysyms = [code for _, _, src, code in INDEX_DEFS if src == "y"]
    px = yf.download(ysyms, start="2024-06-01", interval="1d", auto_adjust=True,
                     group_by="ticker", threads=True, progress=False)
    recs = []
    for country, name, src, code in INDEX_DEFS:
        try:
            if src == "y":
                c = px[code]["Close"].dropna()
            else:
                c = cn_close_series(code)
                time.sleep(0.8)
        except Exception as e:
            log(f"지수 {name} 실패(빈칸): {type(e).__name__}")
            c = pd.Series(dtype=float, index=pd.DatetimeIndex([]))
        r = {k: None for k in INDEX_COLS}
        r["국가"], r["지수"] = country, name
        if len(c):
            ret = lambda n: round(float(c.iloc[-1] / c.iloc[-1 - n] - 1) * 100, 2) if len(c) > n else None
            r["지수레벨"] = round(float(c.iloc[-1]), 2)
            r["1D"], r["1주"], r["1M"], r["3M"] = ret(1), ret(5), ret(21), ret(63)
            base = c[c.index.year == c.index[-1].year - 1]  # 전년 말 종가
            r["YTD"] = round(float(c.iloc[-1] / base.iloc[-1] - 1) * 100, 2) if len(base) else None
            w52 = c[c.index >= c.index[-1] - pd.Timedelta(days=365)]
            lo, hi = float(w52.min()), float(w52.max())
            r["52주위치%"] = round((c.iloc[-1] - lo) / (hi - lo) * 100) if hi > lo else None
            r["기준일"] = c.index[-1].strftime("%Y-%m-%d")
        recs.append(r)
    return pd.DataFrame(recs, columns=INDEX_COLS)


# ---------------------------------------------------------------- 섹터동향
def build_sector_summary(per_market, uni_sect):
    """시장×섹터별 신고/신저 카운트(순강도→강세/약세) + 섹터 밸류에이션(유니버스 중앙값).
    신고/신저 없는 섹터도 표시 — 장세 파악용 전체 섹터 지도."""
    med = lambda v: round(float(np.median(v)), 1) if v else None
    # 12MF는 애널리스트 추정치가 있는 종목만 잡혀 커버리지가 시장별로 크게 다르다
    # (2026-07-31 실측: 미국 94.8% / 일본 75.5% / 홍콩 65.2% / 중국A 34.4%).
    # 표본이 너무 얇으면 섹터 중앙값이 1~2종목 값이 돼 오도하므로 아래 최소 표본 미만은 비운다.
    MIN_FWD_N = 5
    rows = []
    for m in MARKET_ORDER:
        d = per_market.get(m)
        if d is None:
            continue
        stats = uni_sect.get(m, {})
        secs = sorted(set(stats) | (set(d["섹터"]) if len(d) else set()))
        for s in secs:
            st = stats.get(s) or {}
            nh = int(((d["섹터"] == s) & (d["구분"] == "신고가")).sum()) if len(d) else 0
            nl = int(((d["섹터"] == s) & (d["구분"] == "신저가")).sum()) if len(d) else 0
            net = nh - nl
            rows.append({"시장": m, "섹터": s, "유니버스종목수": int(st.get("n", 0)),
                         "신고가": nh, "신저가": nl, "순강도(신고-신저)": net,
                         "판정": "-" if nh == 0 and nl == 0 else
                                 ("강세" if net > 0 else "약세" if net < 0 else "혼조"),
                         "PER(12MF중앙값)": (med(st.get("pe"))
                                          if len(st.get("pe") or []) >= MIN_FWD_N else None),
                         "12MF표본수": len(st.get("pe") or []),
                         "PER(TTM중앙값)": med(st.get("pe_ttm")),
                         "PBR(중앙값)": med(st.get("pb"))})
    df = pd.DataFrame(rows, columns=["시장", "섹터", "유니버스종목수", "신고가", "신저가",
                                     "순강도(신고-신저)", "판정", "PER(12MF중앙값)",
                                     "12MF표본수", "PER(TTM중앙값)", "PBR(중앙값)"])
    if len(df):
        df["_m"] = pd.Categorical(df["시장"], categories=MARKET_ORDER, ordered=True)
        df["_flat"] = df["순강도(신고-신저)"].eq(0)   # 순강도 0(혼조/무신호) → 맨 아래로
        df = (df.sort_values(["_m", "_flat", "순강도(신고-신저)"],
                             ascending=[True, True, False])
              .drop(columns=["_m", "_flat"]).reset_index(drop=True))
    return df


# ---------------------------------------------------------------- 출력
def load_comment(out_dir, tag):
    """당일 코멘트(comment_YYYYMMDD.txt, 스케줄 세션의 Claude가 작성) — 없으면 None."""
    p = out_dir / f"comment_{tag}.txt"
    return p.read_text(encoding="utf-8").strip() if p.exists() else None


def attach_etf_comment(etf_df, out_dir, tag):
    """섹터ETF 표에 '코멘트' 컬럼(섹터 바로 뒤)을 붙인다.
    소스: sector_comment_YYYYMMDD.json = {ETF티커: 그날 섹터 코멘트} (티커는 norm_tk로 매칭).
    json이 있으면 그걸로 코멘트 컬럼을 (재)구성(json에 없는 ETF는 빈칸),
    json이 없고 기존 df에 코멘트가 이미 있으면 그대로 두고 위치만 정렬한다."""
    if etf_df is None or not len(etf_df):
        return etf_df
    p = out_dir / f"sector_comment_{tag}.json"
    if p.exists():
        cm = {norm_tk(k): v for k, v in
              json.loads(p.read_text(encoding="utf-8")).items()}
        etf_df = etf_df.copy()
        etf_df["코멘트"] = [cm.get(norm_tk(t), "") for t in etf_df["티커"]]
    if "코멘트" not in etf_df.columns:
        return etf_df
    cols = [c for c in etf_df.columns if c != "코멘트"]
    ins = cols.index("섹터") + 1 if "섹터" in cols else len(cols)
    return etf_df[cols[:ins] + ["코멘트"] + cols[ins:]]


def write_xlsx(path, per_market, data_dates, note=None, sector_df=None, etf_df=None,
               comment=None, index_df=None):
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    with pd.ExcelWriter(path, engine="openpyxl") as w:
        summary = pd.DataFrame([
            {"시장": m, "데이터 기준일": data_dates.get(m, "-"),
             "신고가": int((d["구분"] == "신고가").sum()), "신저가": int((d["구분"] == "신저가").sum()),
             "그중 52주 신고가": int(((d["구분"] == "신고가") & (d["52주"] == "O")).sum()),
             "그중 52주 신저가": int(((d["구분"] == "신저가") & (d["52주"] == "O")).sum()),
             "오늘 신규진입(NEW)": int((d.get("NEW/OLD", pd.Series(dtype=object)) == "NEW").sum())}
            for m, d in per_market.items()])
        summary.to_excel(w, sheet_name="요약", index=False)
        if index_df is not None and len(index_df):
            index_df.to_excel(w, sheet_name="주요지수", index=False)
        if sector_df is not None and len(sector_df):
            sector_df.to_excel(w, sheet_name="섹터동향", index=False)
        if etf_df is not None and len(etf_df):
            etf_df.to_excel(w, sheet_name="섹터ETF", index=False)
        for m, d in per_market.items():
            # 엑셀 반출은 시총($B) 내림차순 정렬 (2026-07-30 사용자 확정)
            d_x = (d.sort_values("시총($B)", ascending=False, na_position="last")
                   if len(d) and "시총($B)" in d.columns else d)
            d_x.to_excel(w, sheet_name=m, index=False)
        wb = w.book
        r0 = len(per_market) + 3
        if comment:
            from openpyxl.styles import Font as _F
            for i, line in enumerate(comment.split("\n")):  # 여러 줄 코멘트는 행 단위로
                c = wb["요약"].cell(row=r0 + i, column=1,
                                    value=("💬 " + line) if i == 0 else line)
                c.font = _F(bold=(i == 0))
            r0 += comment.count("\n") + 2
        if note:
            wb["요약"].cell(row=r0, column=1, value="※ " + note)
        widths = {"구분": 8, "NEW/OLD": 9, "티커": 19, "종목명": 34, "섹터": 18, "종가": 10,
                  "등락률(%)": 10, "시총($B)": 9, "PER(12MF)": 10, "PER(TTM)": 9, "PBR": 7,
                  "60일": 6, "52주": 6, "이유": 80, "코멘트": 95,
                  "유니버스종목수": 13, "신고가": 8, "신저가": 8, "순강도(신고-신저)": 15, "판정": 8,
                  "PER(12MF중앙값)": 14, "12MF표본수": 10, "PER(TTM중앙값)": 13, "PBR(중앙값)": 10,
                  "국가": 8, "지수": 12, "지수레벨": 12, "YTD": 9, "52주위치%": 10, "기준일": 12}
        up = Font(color="C00000")
        down = Font(color="0000C0")
        for ws in wb.worksheets:
            ws.freeze_panes = "A2"
            hdr = [c.value for c in ws[1]]
            for c in ws[1]:
                c.font = Font(bold=True)
                c.fill = PatternFill("solid", fgColor="DDDDDD")
            for i, h in enumerate(hdr, 1):
                h = str(h)
                default = 8 if (h in ("1D", "1주", "1M", "3M") or h.startswith("YTD")) \
                    else 6 if h.startswith("순위") else 14
                ws.column_dimensions[get_column_letter(i)].width = widths.get(h, default)
            if ws.title not in ("요약", "섹터ETF", "주요지수"):
                ws.auto_filter.ref = ws.dimensions
            if ws.title == "주요지수":
                for ci, h in enumerate(hdr, 1):
                    if str(h) not in ("1D", "1주", "1M", "3M", "YTD"):
                        continue
                    for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                        v = row[0].value
                        row[0].number_format = '+0.00"%";-0.00"%"'
                        if isinstance(v, (int, float)):
                            row[0].font = up if v > 0 else down if v < 0 else Font()
            if ws.title == "섹터ETF":
                # 스크린샷 스타일 히트맵: 컬럼별 상대 색조 (파랑=약세 ~ 빨강=강세)
                for ci, h in enumerate(hdr, 1):
                    h = str(h)
                    if not (h in ("1D", "1주", "1M", "3M") or h.startswith("YTD")):
                        continue
                    col = get_column_letter(ci)
                    vals = [row[0].value for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci)
                            if isinstance(row[0].value, (int, float))]
                    if not vals:
                        continue
                    vmin, vmax = min(vals), max(vals)
                    # 부호 기준 색: 음수=파랑, 0=흰색, 양수=빨강 (컬럼 상대 스케일 아님)
                    if vmin < 0 < vmax:
                        rule = ColorScaleRule(start_type="num", start_value=vmin, start_color="7A9BFF",
                                              mid_type="num", mid_value=0, mid_color="FFFFFF",
                                              end_type="num", end_value=vmax, end_color="FF7B7B")
                    elif vmin >= 0:
                        rule = ColorScaleRule(start_type="num", start_value=0, start_color="FFFFFF",
                                              end_type="num", end_value=max(vmax, 1e-9), end_color="FF7B7B")
                    else:
                        rule = ColorScaleRule(start_type="num", start_value=vmin, start_color="7A9BFF",
                                              end_type="num", end_value=0, end_color="FFFFFF")
                    ws.conditional_formatting.add(f"{col}2:{col}{ws.max_row}", rule)
                    for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                        row[0].number_format = '0.0"%"'
                if "코멘트" in hdr:
                    ci = hdr.index("코멘트") + 1
                    for row in ws.iter_rows(min_row=2, min_col=ci, max_col=ci):
                        row[0].alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=ws.max_row + 2, column=1, value="※ " + ETF_NOTE)
            if "구분" in hdr:
                gi = hdr.index("구분")
                ri = hdr.index("이유")
                ni = hdr.index("NEW/OLD") if "NEW/OLD" in hdr else None
                for row in ws.iter_rows(min_row=2):
                    row[gi].font = up if row[gi].value == "신고가" else down
                    row[ri].alignment = Alignment(wrap_text=True)
                    if ni is not None and row[ni].value == "NEW":
                        row[ni].font = Font(bold=True)
            if "판정" in hdr:
                pi = hdr.index("판정")
                for row in ws.iter_rows(min_row=2):
                    if row[pi].value == "강세":
                        row[pi].font = up
                    elif row[pi].value == "약세":
                        row[pi].font = down
                ws.cell(row=ws.max_row + 2, column=1,
                        value="※ PER(12MF)=선행 컨센서스 기준, 12MF표본수=그 중앙값을 만든 종목 수"
                              " (표본 5개 미만은 오도 방지를 위해 공란). 중국A는 선행 추정 커버리지가"
                              " 34% 수준이라 TTM 중앙값을 함께 보고 판단할 것."
                              " PBR=최근분기 순자산 — 모두 섹터 유니버스($10B+) 중앙값, 적자 제외")


def write_outputs(out_dir, tag, per_market, data_dates, note=None, uni_sect=None, etf_df=None,
                  index_df=None):
    sector_df = build_sector_summary(per_market, uni_sect or {})
    sector_df.to_csv(out_dir / f"sector_{tag}.csv", index=False, encoding="utf-8-sig")
    if etf_df is not None:
        etf_df = attach_etf_comment(etf_df, out_dir, tag)
        etf_df.to_csv(out_dir / f"etf_{tag}.csv", index=False, encoding="utf-8-sig")
    if index_df is not None and len(index_df):
        index_df.to_csv(out_dir / f"index_{tag}.csv", index=False, encoding="utf-8-sig")
    xlsx = out_dir / f"신고신저가_{tag}.xlsx"
    try:
        write_xlsx(xlsx, per_market, data_dates, note, sector_df, etf_df,
                   load_comment(out_dir, tag), index_df)
    except PermissionError:
        log(f"경고: {xlsx.name} 이 열려 있어 xlsx 저장 실패 — 파일을 닫고 재실행 필요")
    # 이유 병합용 스캔 원본 + 뉴스검색 대상 top movers
    all_df = pd.concat([d.assign(시장=m) for m, d in per_market.items()], ignore_index=True)
    all_df.to_csv(out_dir / f"scan_{tag}.csv", index=False, encoding="utf-8-sig")
    # 어제 52주 리스트 (OLD 종목의 '오늘 신규 52주 돌파' 판정용)
    prev = load_prev_scan(out_dir, tag)
    prev52 = set()
    if prev is not None and "52주" in prev.columns:
        for _, pr in prev.iterrows():
            if str(pr.get("52주")).strip() == "O":
                prev52.add((pr["시장"], norm_tk(pr["티커"]), pr["구분"]))
    movers = []
    for m, d in per_market.items():
        for g in ("신고가", "신저가"):
            def row(r, event):
                return {"시장": m, "구분": g, "NEW/OLD": r.get("NEW/OLD", "NEW"),
                        "티커": r["티커"], "종목명": r["종목명"],
                        "등락률(%)": r["등락률(%)"], "event": event}
            # 1) NEW: 신규 진입 상위 (기본 검색 대상)
            new_sub = d[(d["구분"] == g) & (d["NEW/OLD"] == "NEW")].head(TOP_N)
            movers += [row(r, "신규") for _, r in new_sub.iterrows()]
            # 2) OLD인데 오늘 유의미 이벤트: 큰 변동(|등락률|≥7%) 또는 신규 52주 돌파 → 이유 재검색
            od = d[(d["구분"] == g) & (d["NEW/OLD"] == "OLD")].copy()
            if len(od):
                new52 = [str(f).strip() == "O" and (m, norm_tk(t), g) not in prev52
                         for t, f in zip(od["티커"], od["52주"])]
                od["_new52"] = new52
                od["_abs"] = od["등락률(%)"].abs()
                hot = od[(od["_abs"] >= OLD_EVENT_CHG) | od["_new52"]] \
                    .sort_values("_abs", ascending=False).head(OLD_TOP_N)
                movers += [row(r, "52주돌파" if r["_new52"] else "급변") for _, r in hot.iterrows()]
    (out_dir / f"top_movers_{tag}.json").write_text(
        json.dumps({"data_dates": data_dates, "note": note, "movers": movers},
                   ensure_ascii=False, indent=1), encoding="utf-8")
    write_html(out_dir, tag, per_market, data_dates, note)
    return xlsx


def num_or_none(v):
    return None if v is None or (isinstance(v, float) and pd.isna(v)) else float(v)


# ---------------------------------------------------------------- HTML 대시보드
HTML_TMPL = r"""<!doctype html><html lang="ko"><head><meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>신고·신저가 __TAG_DASH__</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/gh/orioncactus/pretendard@v1.3.9/dist/web/variable/pretendardvariable-dynamic-subset.min.css">
<style>
:root{--bg:#0f0f12;--card:#16161a;--line:#232329;--tx:#e8e8ea;--dim:#86868f;
--up:#e25c50;--dn:#6d8fd4;--acc:#ededf0;--tbh:56px}
*{box-sizing:border-box}
html{scrollbar-color:#2e2e36 var(--bg)}
body{margin:0;color:var(--tx);padding:22px 26px 60px;background:var(--bg);
font:14px/1.55 'Pretendard Variable',Pretendard,'Segoe UI','Malgun Gothic',sans-serif;
-webkit-font-smoothing:antialiased}
h1{font-size:20px;margin:0;font-weight:700;letter-spacing:-.3px}
h1 .gr{color:#fff}
.sub{color:var(--dim);font-size:12px;margin-top:3px;letter-spacing:.1px}
header{display:flex;justify-content:space-between;align-items:flex-end;gap:12px;flex-wrap:wrap;margin-bottom:16px}
.actions{display:flex;gap:8px;align-items:center}
.btn{background:var(--acc);color:#101013;font-weight:700;padding:9px 16px;border-radius:9px;
text-decoration:none;font-size:13px;border:none;cursor:pointer;transition:.12s}
.btn:hover{background:#fff}
select{background:var(--card);color:var(--tx);border:1px solid var(--line);border-radius:9px;padding:8px 10px;font:inherit}
#cmt{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 17px;margin-bottom:14px;font-size:13.5px}
.cline{display:flex;gap:11px;margin:6px 0;align-items:flex-start;line-height:1.6}
.cline:first-child span:last-child{font-weight:600;color:#f2f2f4}
.clab{flex:0 0 44px;text-align:center;font-size:11px;font-weight:700;padding:2px 0;border-radius:6px;margin-top:2px;letter-spacing:.4px}
.cards{display:grid;grid-template-columns:repeat(auto-fit,minmax(215px,1fr));gap:10px;margin-bottom:16px}
.card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:13px 16px 12px;
cursor:pointer;transition:border-color .15s,transform .15s;position:relative}
.card:hover{border-color:#3a3a42;transform:translateY(-1px)}
.card.on{border-color:#8b8b95}
.card h3{margin:0 0 8px;font-size:11.5px;color:var(--dim);font-weight:700;letter-spacing:1.2px}
.big{display:flex;align-items:baseline;gap:7px;font-variant-numeric:tabular-nums}
.big b{font-size:26px;font-weight:800;letter-spacing:-.5px;line-height:1}
.big b.u{color:var(--up)}.big b.d{color:var(--dn)}
.big em{font-style:normal;font-size:11px;color:var(--dim);margin-right:9px}
.nw{margin-top:5px;font-size:11px;color:var(--dim)}
.nw b{color:#c9c9cf;font-weight:700}
.spk{position:absolute;right:15px;top:14px;opacity:.9}
.card .bar{height:4px;border-radius:99px;background:#26262c;margin-top:11px;overflow:hidden}
.card .bar i{display:block;height:100%;border-radius:99px;background:var(--up);transition:width .3s}
.toolbar{position:sticky;top:0;z-index:30;background:var(--bg);
padding:10px 0;margin-bottom:10px;border-bottom:1px solid var(--line)}
.controls{display:flex;gap:7px;flex-wrap:wrap;align-items:center}
.chip{background:var(--card);border:1px solid var(--line);color:var(--dim);padding:6.5px 13px;
border-radius:999px;cursor:pointer;font-size:12.5px;user-select:none;transition:.12s;font-weight:600;white-space:nowrap}
.chip:hover{color:var(--tx);border-color:#3a3a42}
.chip.on{color:#101013;border-color:transparent;background:var(--acc);font-weight:700}
.sep{width:1px;height:18px;background:var(--line);margin:0 3px}
input[type=search]{flex:1;min-width:150px;background:var(--card);border:1px solid var(--line);
color:var(--tx);border-radius:999px;padding:8px 15px;outline:none;font:inherit;transition:border-color .15s}
input[type=search]:focus{border-color:#4a4a55}
.view{border-radius:12px}
.anim{animation:fin .18s ease}
@keyframes fin{from{opacity:0}to{opacity:1}}
table{width:100%;border-collapse:separate;border-spacing:0;background:var(--card);border-radius:12px;border:1px solid var(--line)}
thead tr th:first-child{border-top-left-radius:12px}thead tr th:last-child{border-top-right-radius:12px}
th{position:sticky;top:var(--tbh);z-index:20;background:#19191e;color:var(--dim);font-size:11.5px;text-align:left;
padding:10px 11px;cursor:pointer;white-space:nowrap;user-select:none;font-weight:600;letter-spacing:.4px}
th:hover{color:#c9c9cf}
th.sa:after{content:" ↑";color:#c9c9cf}
th.sd:after{content:" ↓";color:#c9c9cf}
td{padding:8.5px 11px;border-top:1px solid #1e1e24;vertical-align:top;white-space:nowrap;font-size:13.5px}
td.nm,td.why{white-space:normal;min-width:150px}
tr:hover td{background:#1c1c22}
.tk{font-family:'Cascadia Code',Consolas,monospace;font-size:12.5px;color:#b9bfca;cursor:pointer;white-space:nowrap;font-weight:600}
.tk:hover{color:#fff}
.up{color:var(--up)}.dn{color:var(--dn)}
.bN{background:var(--acc);color:#101013;font-size:10.5px;font-weight:800;padding:2.5px 8px;border-radius:6px}
.bO{background:#232329;color:var(--dim);font-size:10.5px;padding:2.5px 8px;border-radius:6px}
.bH{background:#3a2a12;color:#ffb867;font-size:10.5px;font-weight:800;padding:2.5px 7px;border-radius:6px;border:1px solid #6b4a1e}
.b52{background:#232329;color:#a9a1c9;font-size:10.5px;padding:2px 6px;border-radius:5px;margin-left:4px;font-weight:700}
.why{color:#b8b8c0;font-size:12.5px}.why .py{color:var(--dim)}
.num{text-align:right;font-variant-numeric:tabular-nums;white-space:nowrap}
.empty{text-align:center;color:var(--dim);padding:30px 0 !important}
footer{color:var(--dim);font-size:11.5px;margin-top:14px}
#toast{position:fixed;bottom:24px;left:50%;transform:translateX(-50%);background:var(--acc);
color:#101013;font-weight:700;padding:10px 20px;border-radius:999px;opacity:0;transition:.25s;pointer-events:none;z-index:99}
#cnt{color:var(--dim);font-size:12px;margin-left:auto;white-space:nowrap}
@media(max-width:680px){
body{padding:12px 12px 44px}
h1{font-size:16.5px}.sub{font-size:10.5px}
.cards{grid-template-columns:1fr 1fr;gap:8px}
.card{padding:11px 12px 10px}
.big b{font-size:20px}.big em{margin-right:5px}
.spk{display:none}
.clab{flex-basis:38px;font-size:10px}
#cmt{font-size:12.5px;padding:11px 12px}
.toolbar{position:static;padding:6px 0}
th{position:static}
.view{overflow-x:auto}
.hm{display:none}
td{font-size:12.5px;padding:7.5px 8px}
.btn{padding:8px 12px;font-size:12px}
}
</style></head><body>
<header>
 <div><h1><span class="gr">신고·신저가 스캔</span> <span style="color:var(--dim);font-weight:500;font-size:14px">__TAG_DASH__</span></h1>
 <div class="sub">기준일 __DATES_LINE__ · 시총 $10B+ · 60일/52주</div></div>
 <div class="actions">
  <span class="chip" id="helpBtn" style="cursor:pointer">ⓘ 도움말</span>
  <select id="dateSel"></select>
  <a class="btn" id="xlsxBtn" href="신고신저가___TAG__.xlsx" download>⬇ 엑셀 반출</a>
 </div>
</header>
<div id="help" style="display:none;background:var(--card);border:1px solid var(--line);border-radius:12px;
padding:13px 17px;margin-bottom:12px;font-size:13px;line-height:1.85">
 <b style="color:var(--acc)">📖 사용법</b>
 <div>· <b>섹터동향</b> 탭에서 섹터를 클릭하면 그 섹터의 <b>신고·신저 종목</b>이 종목 탭에 뜹니다. (필터칩 ✕로 해제)</div>
 <div>· <b>지수</b> 탭 = 한·미·일·중·홍콩 주요지수 등락, <b>섹터ETF</b> 탭 = 섹터별 상대강도 히트맵.</div>
 <div>· 티커 클릭 → 블룸버그 티커 복사 · 표 머리글 클릭 → 정렬 · 검색창 → 티커·종목·섹터·이유 검색.</div>
 <div>· 상단 필터: 시장 / 신고·신저 / NEW만(당일 신규 진입) / 52주만. 우상단 날짜로 과거 스캔·전체기간 조회.</div>
 <div>· 태그: <b style="color:#ededf0">NEW</b> 오늘 신규 진입 · <span style="color:var(--dim)">OLD</span> 연속 등재 · <b style="color:#ffb867">OLD⚡</b> 연속이지만 오늘 새 이벤트(±7%↑ 또는 52주 돌파)로 이유 갱신됨.</div>
 <div>· 💬 코멘트는 개조식 국가별 요약(당일 미작성 시 최근 코멘트 자동 표시). 헤더의 시장별 기준일로 데이터 날짜 확인.</div>
</div>
<div id="datewarn" style="display:none;background:#3a1414;border:1px solid #7a2a2a;color:#ff9b9b;
border-radius:10px;padding:10px 14px;margin-bottom:12px;font-size:13px;font-weight:600"></div>
<div id="cmt" style="display:none"></div>
<div class="cards" id="cards"></div>
<div class="toolbar"><div class="controls">
 <span class="chip vw on" data-v="종목">종목</span><span class="chip vw" data-v="지수">지수</span>
 <span class="chip vw" data-v="섹터동향">섹터동향</span>
 <span class="chip vw" data-v="섹터ETF">섹터ETF</span>
 <span class="sep"></span>
 <span class="chip mkt on" data-v="전체">전체 시장</span><span class="chip mkt" data-v="미국">미국</span>
 <span class="chip mkt" data-v="일본">일본</span>
 <span class="chip mkt" data-v="홍콩">홍콩</span><span class="chip mkt" data-v="중국A">중국A</span>
 <span class="sep"></span>
 <span class="chip gb on" data-v="전체">신고+신저</span><span class="chip gb" data-v="신고가">신고가</span>
 <span class="chip gb" data-v="신저가">신저가</span>
 <span class="chip" id="newOnly">NEW만</span>
 <span class="chip" id="w52Only">52주만</span>
 <span class="chip" id="secpill" style="display:none;background:#26262f;color:var(--acc);border-color:#3a3a42"></span>
 <input type="search" id="q" placeholder="티커 · 종목명 · 섹터 · 이유 검색">
 <span id="cnt"></span>
</div></div>
<div id="vStock" class="view"><table><thead><tr>
<th data-k="dt" id="thDt" style="display:none">날짜</th>
<th data-k="m">시장</th><th data-k="g">구분</th><th data-k="n">NEW</th><th data-k="t">티커</th>
<th data-k="nm">종목명</th><th data-k="s" class="hm">섹터</th><th class="num" data-k="c">종가</th>
<th class="num" data-k="chg">등락률%</th><th class="num hm" data-k="mc">시총$B</th>
<th class="num hm" data-k="pe">PER</th><th class="num hm" data-k="pb">PBR</th><th data-k="why">이유</th>
</tr></thead><tbody id="tb"></tbody></table></div>
<div id="vSector" class="view" style="display:none"></div>
<div id="vEtf" class="view" style="display:none"></div>
<div id="vIndex" class="view" style="display:none"></div>
<footer>__NOTE__ · PER=TTM · 티커 클릭 = 블룸버그 티커 복사 · 생성 __TAG_DASH__</footer>
<div id="toast">복사됨</div>
<script>
const DATA=__DATA__;const DATES=__DATES__;const TAG="__TAG__";
const SDATA=__SDATA__;const EDATA=__EDATA__;const ETFCOLS=__ETFCOLS__;const CMTS=__CMTS__;
const IDATA=__IDATA__;
let st={date:TAG,mkt:"전체",gb:"전체",nw:false,w52:false,q:"",k:"chg",dir:-1,view:"종목",sec:null};
const dash=d=>d.replace(/(\d{4})(\d{2})(\d{2})/,"$1-$2-$3");
const sel=document.getElementById("dateSel");
DATES.slice().reverse().forEach(d=>{const o=document.createElement("option");o.value=d;
o.textContent=dash(d);sel.appendChild(o)});
const oa=document.createElement("option");oa.value="ALL";oa.textContent="전체 기간 (통합검색)";sel.appendChild(oa);
sel.value=TAG;
sel.onchange=()=>{st.date=sel.value;
document.getElementById("xlsxBtn").href="신고신저가_"+(st.date==="ALL"?TAG:st.date)+".xlsx";render()};
function refDate(){return st.date==="ALL"?TAG:st.date}
function fmt(v,d){return v==null||isNaN(v)?"-":Number(v).toLocaleString("en-US",{minimumFractionDigits:d,maximumFractionDigits:d})}
const esc2=s=>s.replace(/&/g,"&amp;").replace(/</g,"&lt;").replace(/>/g,"&gt;");
function setTbh(){document.documentElement.style.setProperty("--tbh",
(document.querySelector(".toolbar").offsetHeight)+"px")}
addEventListener("resize",setTbh);addEventListener("load",setTbh);
if(document.fonts&&document.fonts.ready)document.fonts.ready.then(setTbh);
// ---- 코멘트 (■ 라벨: → 배지)
const CLAB={"결론":"#ffffff","美":"#cfae95","미국":"#cfae95","日":"#c9a3af","일본":"#c9a3af",
"홍콩":"#aaa2cc","中":"#c2bb9a","중국":"#c2bb9a","中·홍콩":"#c2bb9a","좋은 섹터":"#e25c50",
"나쁜 섹터":"#6d8fd4","이벤트":"#b8b8c0","체크":"#86868f"};
function updCmt(){const el=document.getElementById("cmt");
let d=refDate(),t=CMTS[d]||"",carried=false;
if(!t){const ks=Object.keys(CMTS).filter(k=>k<=d).sort();if(ks.length){d=ks[ks.length-1];t=CMTS[d];carried=true}}
if(!t){el.style.display="none";return}
const hdr=carried?`<div class="cline"><span class="clab" style="color:#86868f;border:1px solid #2c2c33">이전</span><span style="color:var(--dim)">${dash(d)} 코멘트 · 당일 미작성</span></div>`:"";
el.innerHTML=hdr+t.split("\n").map(line=>{
const m=line.match(/^■\s*([^:：]+)[:：]\s*(.*)$/);
if(!m)return`<div class="cline">${esc2(line)}</div>`;
const lab=m[1].trim(),c=CLAB[lab]||"#c9c9cf";
return`<div class="cline"><span class="clab" style="color:${c};border:1px solid #2c2c33">${esc2(lab)}</span><span>${esc2(m[2])}</span></div>`}).join("");
el.style.display=""}
// ---- 카드 (빅넘버 + 최근 추이 스파크 + 비율바)
function spark(m){const ds=DATES.slice(-10);if(ds.length<2)return"";
const nets=ds.map(d=>{let h=0,l=0;for(const x of DATA)if(x.m===m&&x.dt===d)x.g==="신고가"?h++:l++;return h-l});
const M=Math.max(1,...nets.map(v=>Math.abs(v)));const W=5,G=3,H=30,mid=H/2;
return`<svg class="spk" width="${ds.length*(W+G)-G}" height="${H}" aria-hidden="true">`+
`<line x1="0" y1="${mid}" x2="${ds.length*(W+G)-G}" y2="${mid}" stroke="#26262c" stroke-width="1"/>`+
nets.map((v,i)=>{const bh=Math.max(1.5,Math.abs(v)/M*(mid-1));
return`<rect x="${i*(W+G)}" y="${v>=0?mid-bh:mid}" width="${W}" height="${bh}" rx="1.2" fill="${v>=0?"var(--up)":"var(--dn)"}" opacity="${i===ds.length-1?1:.45}"/>`}).join("")+`</svg>`}
function cards(){const el=document.getElementById("cards");el.innerHTML="";
const ref=refDate();
["미국","일본","홍콩","중국A"].forEach(m=>{const r=DATA.filter(x=>x.m===m&&x.dt===ref);
const h=r.filter(x=>x.g==="신고가"),l=r.filter(x=>x.g==="신저가");
const hn=h.filter(x=>x.n==="NEW").length,ln=l.filter(x=>x.n==="NEW").length;
const tot=h.length+l.length,p=tot?Math.round(h.length/tot*100):50;
const c=document.createElement("div");c.className="card"+(st.mkt===m?" on":"");
c.innerHTML=`<h3>${m}</h3>${spark(m)}
<div class="big"><b class="u">${h.length}</b><em>신고</em><b class="d">${l.length}</b><em>신저</em></div>
<div class="nw">NEW <b>${hn}</b> 신고 · <b>${ln}</b> 신저</div>
<div class="bar" style="background:${tot?"#33333f":"#26262c"}"><i style="width:${p}%${tot?"":";opacity:.25"}"></i></div>`;
c.onclick=()=>{st.mkt=(st.mkt===m?"전체":m);document.querySelectorAll(".chip.mkt").forEach(x=>
x.classList.toggle("on",x.dataset.v===st.mkt));render()};el.appendChild(c)})}
// ---- 히트맵 (부호 기준: 0=아이보리, 음수=파랑, 양수=빨강)
function heat(v,min,max){if(v==null||isNaN(v))return"";
const M=Math.max(Math.abs(min),Math.abs(max));if(!M)return"";
const t=Math.max(-1,Math.min(1,v/M));let r,g,b;
if(t<0){const u=-t;r=250-138*u;g=248-104*u;b=243-33*u}
else{const u=t;r=250-24*u;g=248-128*u;b=243-133*u}
return`background:rgb(${r|0},${g|0},${b|0});color:#1c1710`}
function bump(el){el.classList.remove("anim");void el.offsetWidth;el.classList.add("anim")}
// ---- 섹터동향 뷰
function renderSector(){let r=SDATA.filter(x=>x.dt===refDate()&&(st.mkt==="전체"||x.m===st.mkt));
const MKO=["미국","일본","홍콩","중국A"];
r=r.slice().sort((a,b)=>{const ma=MKO.indexOf(a.m),mb=MKO.indexOf(b.m);if(ma!==mb)return ma-mb;
const fa=a.net===0?1:0,fb=b.net===0?1:0;if(fa!==fb)return fa-fb;return b.net-a.net});
document.getElementById("cnt").textContent=r.length+"섹터";
const el=document.getElementById("vSector");
el.innerHTML='<table><thead><tr><th>시장</th><th>섹터</th>'+
'<th class="num">유니버스</th><th class="num">신고</th><th class="num">신저</th><th class="num">순강도</th>'+
'<th>판정</th><th class="num">PER(12MF)</th><th class="num">PBR</th></tr></thead><tbody>'+
(r.length?r.map(x=>`<tr style="cursor:pointer" data-s="${esc2(x.s)}" onclick="drillSector('${x.m}',this.dataset.s)" title="클릭 → 이 섹터의 신고·신저 종목 보기">
<td>${x.m}</td><td>${x.s}</td><td class="num">${x.n}</td>
<td class="num up">${x.nh||""}</td><td class="num dn">${x.nl||""}</td>
<td class="num ${x.net>0?"up":x.net<0?"dn":""}">${x.net>0?"+"+x.net:x.net}</td>
<td class="${x.j==="강세"?"up":x.j==="약세"?"dn":""}">${x.j}</td>
<td class="num">${fmt(x.pe,1)}</td><td class="num">${fmt(x.pb,1)}</td></tr>`).join("")
:'<tr><td colspan="9" class="empty">해당 날짜 데이터 없음</td></tr>')+
'</tbody></table><footer>※ 행 클릭 → 그 섹터의 신고·신저 종목 · 순강도순(강세→약세)·순강도 0 무신호 섹터는 아래로 · PER=TTM·PBR=최근분기 — 섹터 유니버스($10B+) 중앙값, 적자 제외</footer>';bump(el)}
// ---- 주요 지수 뷰
function renderIndex(){const CO=["미국","한국","일본","중국","홍콩"];
let r=IDATA.filter(x=>x.dt===refDate());
r=r.slice().sort((a,b)=>CO.indexOf(a.c)-CO.indexOf(b.c));
document.getElementById("cnt").textContent=r.length+"개 지수";
const ICOL=["1D","1주","1M","3M","YTD"];
const mm=ICOL.map((_,i)=>{const vs=r.map(x=>x.v[i]).filter(v=>v!=null&&!isNaN(v));
return vs.length?[Math.min(...vs),Math.max(...vs)]:[0,0]});
const el=document.getElementById("vIndex");
el.innerHTML='<table><thead><tr><th>국가</th><th>지수</th><th class="num">지수레벨</th>'+
ICOL.map(c=>`<th class="num">${c}</th>`).join("")+
'<th class="num hm">52주위치</th><th class="hm">기준일</th></tr></thead><tbody>'+
(r.length?r.map(x=>`<tr><td>${x.c}</td><td style="font-weight:600">${x.nm}</td>
<td class="num">${x.lv==null?"-":x.lv.toLocaleString("en-US",{minimumFractionDigits:2,maximumFractionDigits:2})}</td>`+
x.v.map((v,i)=>`<td class="num" style="${heat(v,mm[i][0],mm[i][1])}">${v==null?"-":(v>=0?"+":"")+fmt(v,2)+"%"}</td>`).join("")+
`<td class="num hm">${x.w52==null?"-":x.w52+"%"}</td><td class="hm" style="color:var(--dim)">${x.asof||"-"}</td></tr>`).join("")
:'<tr><td colspan="10" class="empty">해당 날짜 데이터 없음</td></tr>')+
'</tbody></table><footer>※ 부호 히트맵(파랑=하락·빨강=상승) · 52주위치=52주 레인지 내 현재가 위치(0=최저,100=최고) · 미·한·일=지수, 중·홍=eastmoney · 기준일은 시장별 최근 정규장</footer>';bump(el)}
// ---- 섹터 드릴다운: 섹터동향 행 클릭 → 그 섹터의 신고·신저 종목
function drillSector(m,s){st.sec={m:m,s:s};st.mkt=m;st.gb="전체";st.view="종목";
document.querySelectorAll(".chip.vw").forEach(x=>x.classList.toggle("on",x.dataset.v==="종목"));
document.querySelectorAll(".chip.mkt").forEach(x=>x.classList.toggle("on",x.dataset.v===m));
document.querySelectorAll(".chip.gb").forEach(x=>x.classList.toggle("on",x.dataset.v==="전체"));
render();window.scrollTo({top:0,behavior:"smooth"})}
// ---- 섹터ETF 뷰
function renderEtf(){let r=EDATA.filter(x=>x.dt===refDate()&&(st.mkt==="전체"||x.m===st.mkt));
const i1d=ETFCOLS.indexOf("1D"),MKO=["미국","일본","홍콩","중국A"];
const flat=v=>(v==null||isNaN(v))?1:(Math.abs(v)<1?1:0);
r=r.slice().sort((a,b)=>{const ma=MKO.indexOf(a.m),mb=MKO.indexOf(b.m);if(ma!==mb)return ma-mb;
const fa=flat(a.v[i1d]),fb=flat(b.v[i1d]);if(fa!==fb)return fa-fb;
const av=a.v[i1d],bv=b.v[i1d],ax=(av==null||isNaN(av))?-1e9:av,bx=(bv==null||isNaN(bv))?-1e9:bv;return bx-ax});
document.getElementById("cnt").textContent=r.length+"ETF";
const mm=ETFCOLS.map((_,i)=>{const vs=r.map(x=>x.v[i]).filter(v=>v!=null&&!isNaN(v));
return vs.length?[Math.min(...vs),Math.max(...vs)]:[0,0]});
const el=document.getElementById("vEtf");
el.innerHTML='<table><thead><tr><th>시장</th><th>티커</th><th>섹터</th>'+
ETFCOLS.map(c=>`<th class="num">${c}</th>`).join("")+'</tr></thead><tbody>'+
(r.length?r.map(x=>`<tr><td>${x.m}</td><td class="tk" onclick="cp('${x.t}')">${x.t}</td><td>${x.s}</td>`+
x.v.map((v,i)=>`<td class="num" style="${heat(v,mm[i][0],mm[i][1])}">${v==null?"-":fmt(v,1)+"%"}</td>`).join("")+
"</tr>").join("")
:'<tr><td colspan="'+(ETFCOLS.length+3)+'" class="empty">해당 날짜 데이터 없음</td></tr>')+
'</tbody></table><footer>※ 0=흰색 기준 부호 히트맵(파랑=하락·빨강=상승) · 배당 포함 총수익 · 시장별 1D 변동 1%↑는 위(상승→하락순)·1% 미만 보합은 아래로</footer>';bump(el)}
// ---- 종목 뷰 + 라우팅
function render(){cards();updCmt();setTbh();
document.getElementById("vStock").style.display=st.view==="종목"?"":"none";
document.getElementById("vSector").style.display=st.view==="섹터동향"?"":"none";
document.getElementById("vEtf").style.display=st.view==="섹터ETF"?"":"none";
document.getElementById("vIndex").style.display=st.view==="지수"?"":"none";
// 섹터 드릴 필터 pill (종목 뷰에서만 유효)
const pill=document.getElementById("secpill");
if(st.sec&&st.view==="종목"){pill.textContent="섹터: "+st.sec.s+"  ✕";pill.style.display=""}
else{pill.style.display="none"}
if(st.view==="섹터동향"){renderSector();return}
if(st.view==="섹터ETF"){renderEtf();return}
if(st.view==="지수"){renderIndex();return}
let r=DATA.filter(x=>
(st.date==="ALL"||x.dt===st.date)&&
(st.mkt==="전체"||x.m===st.mkt)&&(st.gb==="전체"||x.g===st.gb)&&(!st.nw||x.n==="NEW")&&(!st.w52||x.d52)&&
(!st.sec||(x.m===st.sec.m&&x.s===st.sec.s)));
if(st.q){const q=st.q.toLowerCase();r=r.filter(x=>(x.t+x.nm+x.s+x.why).toLowerCase().includes(q))}
r.sort((a,b)=>{if(st.date==="ALL"&&a.dt!==b.dt)return b.dt.localeCompare(a.dt);
let A=a[st.k],B=b[st.k];if(A==null)return 1;if(B==null)return -1;
if(typeof A==="string")return st.dir*A.localeCompare(B);return st.dir*(A-B)});
const total=r.length,CAP=1500;if(r.length>CAP)r=r.slice(0,CAP);
document.getElementById("thDt").style.display=st.date==="ALL"?"":"none";
document.getElementById("cnt").textContent=total+"종목"+(total>CAP?` (상위 ${CAP}만 표시)`:"");
document.querySelectorAll("#vStock th").forEach(h=>{h.classList.toggle("sa",h.dataset.k===st.k&&st.dir>0);
h.classList.toggle("sd",h.dataset.k===st.k&&st.dir<0)});
document.getElementById("tb").innerHTML=r.length?r.map(x=>`<tr>
${st.date==="ALL"?`<td style="color:var(--dim)">${dash(x.dt)}</td>`:""}
<td>${x.m}</td><td class="${x.g==="신고가"?"up":"dn"}">${x.g}</td>
<td>${x.n==="NEW"?'<span class="bN">NEW</span>':(x.why&&x.why.trim()&&!x.why.startsWith("(전일)")?'<span class="bH" title="OLD지만 오늘 새 이벤트로 이유 갱신">OLD⚡</span>':'<span class="bO">OLD</span>')}</td>
<td class="tk" onclick="cp('${x.t}')">${x.t}${x.d52?'<span class="b52">52W</span>':''}</td>
<td class="nm">${x.nm}</td><td class="hm">${x.s}</td><td class="num">${fmt(x.c,2)}</td>
<td class="num ${x.chg>=0?"up":"dn"}">${x.chg==null?"-":(x.chg>=0?"+":"")+fmt(x.chg,2)}</td>
<td class="num hm">${fmt(x.mc,1)}</td>
<td class="num hm">${fmt(x.pe,1)}</td><td class="num hm">${fmt(x.pb,1)}</td>
<td class="why">${x.why&&x.why.startsWith("(전일)")?'<span class="py">'+x.why+"</span>":x.why||""}</td></tr>`).join("")
:'<tr><td colspan="13" class="empty">조건에 맞는 종목이 없습니다</td></tr>';
bump(document.getElementById("vStock"))}
function cp(t){const f=()=>{const a=document.createElement("textarea");a.value=t;document.body.appendChild(a);
a.select();document.execCommand("copy");a.remove()};
(navigator.clipboard?navigator.clipboard.writeText(t).catch(f):Promise.resolve(f())).then?.(()=>{});
const o=document.getElementById("toast");o.textContent=t+" 복사됨";o.style.opacity=1;
setTimeout(()=>o.style.opacity=0,1200)}
document.querySelectorAll(".chip.vw").forEach(c=>c.onclick=()=>{st.view=c.dataset.v;st.sec=null;
document.querySelectorAll(".chip.vw").forEach(x=>x.classList.toggle("on",x===c));render()});
document.querySelectorAll(".chip.mkt").forEach(c=>c.onclick=()=>{st.mkt=c.dataset.v;st.sec=null;
document.querySelectorAll(".chip.mkt").forEach(x=>x.classList.toggle("on",x===c));render()});
document.querySelectorAll(".chip.gb").forEach(c=>c.onclick=()=>{st.gb=c.dataset.v;
document.querySelectorAll(".chip.gb").forEach(x=>x.classList.toggle("on",x===c));render()});
document.getElementById("secpill").onclick=()=>{st.sec=null;render()};
document.getElementById("helpBtn").onclick=e=>{const h=document.getElementById("help");
const show=h.style.display==="none";h.style.display=show?"":"none";e.target.classList.toggle("on",show)};
document.getElementById("newOnly").onclick=e=>{st.nw=!st.nw;e.target.classList.toggle("on",st.nw);render()};
document.getElementById("w52Only").onclick=e=>{st.w52=!st.w52;e.target.classList.toggle("on",st.w52);render()};
document.getElementById("q").oninput=e=>{st.q=e.target.value;render()};
document.querySelectorAll("#vStock th").forEach(h=>h.onclick=()=>{const k=h.dataset.k;
if(st.k===k)st.dir*=-1;else{st.k=k;st.dir=k==="chg"||k==="mc"||k==="c"||k==="pe"||k==="pb"?-1:1}render()});
(()=>{const ft=document.querySelector("footer")?.textContent||"";
const i=ft.indexOf("⚠️");if(i<0)return;
const w=document.getElementById("datewarn");
w.textContent="⚠️ "+ft.slice(i+2).split("  —  ")[0].trim();w.style.display=""})();
render();
</script></body></html>"""


HIST_DATES = 90  # index.html에 내장할 최근 스캔 일수


def write_html(out_dir, tag, per_market, data_dates, note):
    """최근 HIST_DATES개 scan CSV를 전부 내장한 단일 대시보드(index.html) 생성 —
    링크 하나로 날짜 전환·전체기간 통합검색."""
    def num(v):
        return None if pd.isna(v) else float(v)

    rows, dates = [], []
    for p in sorted(out_dir.glob("scan_*.csv"))[-HIST_DATES:]:
        d_tag = p.stem.split("_")[-1]
        if not d_tag.isdigit():
            continue
        df = pd.read_csv(p, encoding="utf-8-sig")
        # 2026-07-31 PER을 TTM→12MF로 전환 — 과거 csv는 옛 컬럼명이라 둘 다 받는다.
        pe_col = "PER(12MF)" if "PER(12MF)" in df.columns else "PER(TTM)"
        has_val = pe_col in df.columns
        dates.append(d_tag)
        for _, r in df.iterrows():
            rows.append({"dt": d_tag, "m": r["시장"], "g": r["구분"],
                         "n": r["NEW/OLD"] if isinstance(r["NEW/OLD"], str) else "NEW",
                         "t": r["티커"], "nm": r["종목명"], "s": r["섹터"], "c": num(r["종가"]),
                         "chg": num(r["등락률(%)"]), "mc": num(r["시총($B)"]),
                         "pe": num(r[pe_col]) if has_val else None,
                         "pb": num(r["PBR"]) if has_val else None,
                         "d52": r["52주"] == "O",
                         "why": r["이유"] if isinstance(r["이유"], str) else ""})
    srows = []
    for p in sorted(out_dir.glob("sector_*.csv"))[-HIST_DATES:]:
        d_tag = p.stem.split("_")[-1]
        if not d_tag.isdigit():
            continue
        for _, r in pd.read_csv(p, encoding="utf-8-sig").iterrows():
            srows.append({"dt": d_tag, "m": r["시장"], "s": r["섹터"],
                          "n": int(r["유니버스종목수"]), "nh": int(r["신고가"]), "nl": int(r["신저가"]),
                          "net": int(r["순강도(신고-신저)"]), "j": r["판정"],
                          "pe": num(r.get("PER(12MF중앙값)", r.get("PER(TTM중앙값)"))),
                          "pb": num(r.get("PBR(중앙값)"))})
    erows, etf_cols = [], []
    for p in sorted(out_dir.glob("etf_*.csv"))[-HIST_DATES:]:
        d_tag = p.stem.split("_")[-1]
        if not d_tag.isdigit():
            continue
        edf = pd.read_csv(p, encoding="utf-8-sig")
        etf_cols = [c for c in edf.columns
                    if c in ("1D", "1주", "1M", "3M") or c.startswith("YTD")]
        for _, r in edf.iterrows():
            erows.append({"dt": d_tag, "m": r["시장"], "t": r["티커"], "s": r["섹터"],
                          "v": [num(r[c]) for c in etf_cols]})
    irows, idx_cols = [], ["1D", "1주", "1M", "3M", "YTD"]
    for p in sorted(out_dir.glob("index_*.csv"))[-HIST_DATES:]:
        d_tag = p.stem.split("_")[-1]
        if not d_tag.isdigit():
            continue
        for _, r in pd.read_csv(p, encoding="utf-8-sig").iterrows():
            irows.append({"dt": d_tag, "c": r["국가"], "nm": r["지수"],
                          "lv": num(r["지수레벨"]), "v": [num(r[k]) for k in idx_cols],
                          "w52": num(r["52주위치%"]),
                          "asof": r["기준일"] if isinstance(r["기준일"], str) else ""})
    cmts = {}
    for p in sorted(out_dir.glob("comment_*.txt"))[-HIST_DATES:]:
        d_tag = p.stem.split("_")[-1]
        if d_tag.isdigit():
            cmts[d_tag] = p.read_text(encoding="utf-8").strip()
    dates_line = " / ".join(f"{m} {v}" for m, v in data_dates.items())
    esc = lambda o: json.dumps(o, ensure_ascii=False).replace("<", "\\u003c")
    html = (HTML_TMPL
            .replace("__DATA__", esc(rows))
            .replace("__CMTS__", esc(cmts))
            .replace("__SDATA__", esc(srows))
            .replace("__EDATA__", esc(erows))
            .replace("__ETFCOLS__", esc(etf_cols))
            .replace("__IDATA__", esc(irows))
            .replace("__DATES_LINE__", dates_line)
            .replace("__DATES__", json.dumps(dates))
            .replace("__TAG_DASH__", f"{tag[:4]}-{tag[4:6]}-{tag[6:]}")
            .replace("__TAG__", tag)
            .replace("__NOTE__", note or ""))
    (out_dir / "index.html").write_text(html, encoding="utf-8")
    log(f"HTML 대시보드({len(dates)}일치): {out_dir / 'index.html'}")
    publish_pages(out_dir / "index.html")


PAGES_DIR = Path.home() / "Documents" / "GitHub" / "highlow-pages"  # GitHub Pages 발행용 클론


def publish_pages(index_html):
    """PAGES_DIR가 git 클론이면 index.html + 날짜별 xlsx 복사 후 commit+push (없으면 건너뜀)."""
    if not (PAGES_DIR / ".git").exists():
        return
    import shutil
    import subprocess
    shutil.copyfile(index_html, PAGES_DIR / "index.html")
    for x in index_html.parent.glob("신고신저가_*.xlsx"):  # 대시보드 '엑셀 반출' 버튼용
        dst = PAGES_DIR / x.name
        # 같은 날 재생성(시트 추가·이유 병합)되면 덮어써야 반출본도 최신이 됨
        if not dst.exists() or x.stat().st_mtime > dst.stat().st_mtime:
            shutil.copyfile(x, dst)
    try:
        # 자격증명이 없으면 Git Credential Manager가 GUI 로그인창을 띄운 채 멈춘다(2026-08-13 실측 33분 소요).
        # 무인 실행이므로 대화형 인증 경로를 전부 차단해 즉시 실패시킨다.
        env = {**os.environ, "GIT_TERMINAL_PROMPT": "0", "GCM_INTERACTIVE": "never"}
        noint = ["-c", "credential.interactive=false", "-c", "credential.guiPrompt=false",
                 "-c", "core.askPass="]
        run = lambda *a, t=60: subprocess.run(["git", "-C", str(PAGES_DIR), *noint, *a],
                                              capture_output=True, text=True, timeout=t, env=env)
        run("add", "-A")
        if run("diff", "--cached", "--quiet").returncode != 0:
            run("commit", "-m", f"update {dt.date.today().isoformat()}")
        # 커밋 생성 여부와 무관하게 '밀린 커밋이 있으면' 항상 push한다.
        # (구버전은 새 커밋이 있을 때만 push해서, 한 번 실패하면 백로그가 조용히 쌓였다)
        ahead = run("rev-list", "--count", "@{upstream}..HEAD").stdout.strip()
        if ahead.isdigit() and int(ahead) > 0:
            p = run("push", t=300)
            if p.returncode == 0:
                log(f"Pages 발행 완료(커밋 {ahead}건 push)")
            else:
                log(f"Pages push 실패(미발행 {ahead}건 누적): {p.stderr.strip()[:160]}")
        else:
            log("Pages 발행: 변경 없음")
    except Exception as e:
        log(f"Pages 발행 실패(무시): {type(e).__name__}")


# ---------------------------------------------------------------- main
def scan_legacy(fx, out_dir):
    log("유니버스 수집…")
    us, hk, cn = get_universes(fx, out_dir / "_cache")
    log(f"유니버스: 미국 {len(us)} / 홍콩 {len(hk)} / 중국A {len(cn)}")
    univs = {"미국": us, "홍콩": hk, "중국A": cn}
    jp_cache = out_dir / "_cache" / "universe_jp.csv"
    if jp_cache.exists():  # 일본은 직전 TV 성공분 캐시로만 폴백 가능
        jp = pd.read_csv(jp_cache, encoding="utf-8-sig", dtype=str)
        if "pe_fwd" in jp.columns:  # TV 캐시는 12MF, 폴백 내부 키는 pe_ttm으로 통일
            jp = jp.rename(columns={"pe_fwd": "pe_ttm"})
        for c in ("mcap_usd", "pe_ttm", "pbr"):
            jp[c] = pd.to_numeric(jp[c], errors="coerce")
        univs = {"미국": us, "일본": jp, "홍콩": hk, "중국A": cn}
    else:
        log("일본 유니버스 캐시 없음 → 일본 스킵 (TV 경로가 한 번 성공하면 생김)")
    per_market, data_dates, uni_sect = {}, {}, {}
    for m, univ in univs.items():
        log(f"{m} 일봉 수집…")
        hists = fetch_history(list(univ["yahoo"]), m)
        d, last = build_rows(univ, hists)
        per_market[m] = d
        data_dates[m] = last or "-"
        uni_sect[m] = {
            # 폴백 소스는 선행 컨센서스를 주지 않아 pe가 곧 TTM — 두 키에 같은 값을 넣어
            # 섹터 시트의 TTM 열이 비지 않게 하되, 12MF 열은 표본수 규칙에 따라 그대로 걸러진다.
            s: {"n": len(g),
                "pe": [x for x in pd.to_numeric(g["pe_ttm"], errors="coerce") if x > 0],
                "pe_ttm": [x for x in pd.to_numeric(g["pe_ttm"], errors="coerce") if x > 0],
                "pb": [x for x in pd.to_numeric(g["pbr"], errors="coerce") if x > 0]}
            for s, g in univ.groupby("sector")}
        log(f"{m}: 일봉 {len(hists)}/{len(univ)}")
    if len(per_market["미국"]):
        log("미국 밸류에이션 수집…")
        fill_us_valuation(per_market["미국"])
    return per_market, data_dates, uni_sect


def freshness_check(data_dates, tag):
    """날짜 무결성 검사 — 데이터 기준일이 실행일(tag) 대비 미래이거나 과도하게 오래되면 경고.
    금융 데이터라 날짜 오류는 치명적 → 조용히 넘기지 않고 로그+산출물에 표시."""
    warns = []
    try:
        run = dt.datetime.strptime(tag, "%Y%m%d").date()
    except ValueError:
        return [f"실행일(tag={tag}) 파싱 실패"]
    for m, ds in data_dates.items():
        try:
            dd = dt.datetime.strptime(str(ds), "%Y-%m-%d").date()
        except ValueError:
            warns.append(f"{m} 기준일 없음/파싱불가({ds})")
            continue
        gap = (run - dd).days
        if gap < 0:
            warns.append(f"{m} 기준일 {ds}이 실행일 {run}보다 미래 — 시계/태그 오류 의심")
        elif gap > 5:  # 주말+공휴일 최대치를 넘는 정체 = API가 옛 데이터 반환 의심
            warns.append(f"{m} 기준일 {ds} — 실행일 {run} 대비 {gap}일 정체(데이터 stale 의심)")
    return warns


def run_scan(out_dir, source="tv"):
    fx = get_fx()
    per_market, note = None, TV_NOTE
    if source == "tv":
        try:
            per_market, data_dates, uni_sect = scan_tradingview(fx, out_dir)
        except Exception as e:
            log(f"TradingView 실패 → eastmoney/yfinance 폴백: {type(e).__name__} {e}")
    if per_market is None:
        per_market, data_dates, uni_sect = scan_legacy(fx, out_dir)
        note = LEGACY_NOTE
    for m, d in per_market.items():
        log(f"{m}: 신고가 {(d['구분'] == '신고가').sum()} / 신저가 {(d['구분'] == '신저가').sum()}"
            f" (기준일 {data_dates[m]})")
    tag = dt.date.today().strftime("%Y%m%d")
    warns = freshness_check(data_dates, tag)
    if warns:
        for w in warns:
            log(f"⚠️ 날짜경고: {w}")
        note = "⚠️ 날짜 확인 필요: " + " / ".join(warns) + "  —  " + note
    apply_new_old(per_market, load_prev_scan(out_dir, tag))
    try:
        log("섹터 ETF 수익률 계산…")
        etf_df = build_etf_table()
    except Exception as e:
        log(f"섹터 ETF 실패(건너뜀): {type(e).__name__} {e}")
        etf_df = None
    try:
        log("주요 지수 수집…")
        index_df = build_index_table()
    except Exception as e:
        log(f"주요 지수 실패(건너뜀): {type(e).__name__} {e}")
        index_df = None
    xlsx = write_outputs(out_dir, tag, per_market, data_dates, note, uni_sect, etf_df, index_df)
    log(f"완료: {xlsx}")


def merge_reasons(out_dir, reasons_path, tag):
    raw_reasons = json.loads(Path(reasons_path).read_text(encoding="utf-8"))
    reasons = {norm_tk(k): v for k, v in raw_reasons.items()}
    # 심볼만 적힌 키("TEAM")도 받아준다 — norm_tk는 "TEAM US"라 그대로면 무음 실패.
    for k, v in raw_reasons.items():
        reasons.setdefault(norm_tk(k).split()[0], v)
    csv = out_dir / f"scan_{tag}.csv"
    all_df = pd.read_csv(csv, encoding="utf-8-sig").fillna({"이유": "", "52주": "", "60일": ""})

    def pick(t):
        n = norm_tk(t)
        return reasons.get(n) or reasons.get(n.split()[0])

    all_df["이유"] = all_df["티커"].map(pick).fillna(all_df["이유"])
    hit = sum(1 for t in all_df["티커"] if pick(t))
    print(f"[merge] 이유 매칭 {hit}/{len(raw_reasons)}건", flush=True)
    all_df.to_csv(csv, index=False, encoding="utf-8-sig")  # csv에도 반영 → 익일 OLD 이유 승계
    meta = json.loads((out_dir / f"top_movers_{tag}.json").read_text(encoding="utf-8"))
    dates, note = meta["data_dates"], meta.get("note")
    per_market = {m: d.drop(columns="시장").reset_index(drop=True)
                  for m, d in all_df.groupby("시장", sort=False)}
    per_market = {m: per_market[m] for m in MARKET_ORDER if m in per_market}
    sector_csv = out_dir / f"sector_{tag}.csv"
    sector_df = pd.read_csv(sector_csv, encoding="utf-8-sig") if sector_csv.exists() else None
    etf_csv = out_dir / f"etf_{tag}.csv"
    etf_df = pd.read_csv(etf_csv, encoding="utf-8-sig") if etf_csv.exists() else None
    etf_df = attach_etf_comment(etf_df, out_dir, tag)   # sector_comment json → 코멘트 컬럼
    if etf_df is not None:
        etf_df.to_csv(etf_csv, index=False, encoding="utf-8-sig")  # 코멘트 반영 지속
    index_csv = out_dir / f"index_{tag}.csv"
    index_df = pd.read_csv(index_csv, encoding="utf-8-sig") if index_csv.exists() else None
    xlsx = out_dir / f"신고신저가_{tag}.xlsx"
    try:
        write_xlsx(xlsx, per_market, dates, note, sector_df, etf_df,
                   load_comment(out_dir, tag), index_df)
    except PermissionError:
        log(f"경고: {xlsx.name} 이 열려 있어 xlsx 저장 실패 — 파일을 닫고 재실행 필요")
    write_html(out_dir, tag, per_market, dates, note)
    log(f"이유 병합 완료: {xlsx}")


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--out-dir", default=str(Path.home() / "Desktop" / "신고신저가"))
    ap.add_argument("--merge-reasons")
    ap.add_argument("--date", default=dt.date.today().strftime("%Y%m%d"))
    ap.add_argument("--source", choices=["tv", "legacy"], default="tv",
                    help="tv=TradingView scanner(기본), legacy=eastmoney/yfinance")
    a = ap.parse_args()
    out_dir = Path(a.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    if a.merge_reasons:
        merge_reasons(out_dir, a.merge_reasons, a.date)
    else:
        run_scan(out_dir, a.source)
